from celery import shared_task
from openpyxl.utils import get_column_letter

from .models import Discipline, Student
import openpyxl


@shared_task
def generate_report():
    wb = wb = openpyxl.Workbook()
    sheet_direction = wb['Sheet']
    sheet_direction.title = 'Directions'
    sheet_groups = wb.create_sheet('Groups')
    disciplines_queryset = Discipline.objects.select_related('direction')
    students_queryset = Student.objects.select_related('group').order_by('surname')
    direction_info = {}
    group_info = {}
    for i in disciplines_queryset:
        if i.direction not in direction_info.keys():
            direction_info[i.direction] = []
        direction_info[i.direction].append(i.title)
    row_counter = 1
    column_counter = 1
    for key, value in direction_info.items():
        sheet_direction.cell(row=row_counter, column=column_counter).value = key.title
        sheet_direction.column_dimensions[get_column_letter(column_counter)].width = len(key.title) * 2.5
        row_counter += 1
        for v in value:
            sheet_direction.cell(row=row_counter, column=column_counter).value = v
            row_counter += 1
        sheet_direction.cell(row=row_counter + 1,
                             column=column_counter).value = f"{key.curator.username} {key.curator.email}"
        row_counter = 1
        column_counter += 1

    for i in students_queryset:
        if i.group not in group_info.keys():
            group_info[i.group] = []
        group_info[i.group].append(i)
    row_counter = 1
    column_counter = 1
    current_max_row = 1
    for key, value in group_info.items():
        sheet_groups.cell(row=row_counter, column=column_counter).value = key.title
        sheet_groups.column_dimensions[get_column_letter(column_counter)].width = len(key.title) * 4
        row_counter += 1
        gender_counter = {}

        for v in value:
            sheet_groups.cell(row=row_counter, column=column_counter).value = f"{v.surname} {v.name}"
            row_counter += 1
            if current_max_row < row_counter:
                current_max_row = row_counter
            if v.sex not in gender_counter:
                gender_counter[v.sex] = 0
            gender_counter[v.sex] += 1

        male = gender_counter.get('M', 0)
        female = gender_counter.get('F', 0)
        sheet_groups.cell(row=current_max_row + 1,
                          column=column_counter).value = (f"Male: {male}\n "
                                                          f"Female: {female}")
        total_students_count = male + female
        sheet_groups.cell(row=current_max_row + 2,
                          column=column_counter).value = f"Free places in group: {20 - total_students_count}"

        row_counter = 1
        column_counter += 1
    wb.save(filename="Test.xlsx")
    return True
